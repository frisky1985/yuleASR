#!/usr/bin/env python3
"""
NvM Configuration Generator Tool
================================
Generates NvM_Cfg.h and NvM_Lcfg.c from JSON/YAML/Excel configuration

Author: yuleASR Team
Version: 1.0.0
"""

import json
import sys
from pathlib import Path
from dataclasses import dataclass, asdict
from typing import List, Optional, Dict
from enum import Enum

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

#==============================================================================
# Enums for Configuration
#==============================================================================
class BlockManagementType(str, Enum):
    NATIVE = "NVM_BLOCK_NATIVE"
    REDUNDANT = "NVM_BLOCK_REDUNDANT"
    DATASET = "NVM_BLOCK_DATASET"

class BlockCrcType(str, Enum):
    NONE = "NVM_CRC_NONE"
    CRC8 = "NVM_CRC_8"
    CRC16 = "NVM_CRC_16"
    CRC32 = "NVM_CRC_32"

class StdStatus(str, Enum):
    ON = "STD_ON"
    OFF = "STD_OFF"

#==============================================================================
# Data Classes
#==============================================================================
@dataclass
class NvMBlockDescriptor:
    """NvM Block Descriptor Configuration"""
    BlockId: int
    BlockName: str
    BlockBaseNumber: int
    ManagementType: BlockManagementType
    NumberOfNvBlocks: int
    NumberOfDataSets: int
    NvBlockLength: int
    NvBlockNum: int
    RomBlockNum: int
    CrcType: BlockCrcType
    BlockUseCrc: bool
    BlockUseSetRamBlockStatus: bool
    BlockWriteProt: bool
    BlockWriteOnce: bool
    BlockAutoValidation: bool
    BlockUseMirror: bool
    BlockUseCompression: bool
    Description: str = ""
    InitCallback: str = "NULL_PTR"
    JobEndCallback: str = "NULL_PTR"

@dataclass
class NvMGeneralConfig:
    """NvM General Configuration"""
    DevErrorDetect: StdStatus = StdStatus.ON
    VersionInfoApi: StdStatus = StdStatus.ON
    SetRamBlockStatusApi: StdStatus = StdStatus.ON
    GetErrorStatusApi: StdStatus = StdStatus.ON
    SetBlockProtectionApi: StdStatus = StdStatus.ON
    GetBlockProtectionApi: StdStatus = StdStatus.OFF
    SetDataIndexApi: StdStatus = StdStatus.ON
    GetDataIndexApi: StdStatus = StdStatus.OFF
    CancelJobApi: StdStatus = StdStatus.ON
    KillWriteAllApi: StdStatus = StdStatus.OFF
    KillReadAllApi: StdStatus = StdStatus.OFF
    RepairDamagedBlocksApi: StdStatus = StdStatus.OFF
    CalcRamBlockCrc: StdStatus = StdStatus.ON
    UseCrcCompMechanism: StdStatus = StdStatus.ON
    MaxNumberOfWriteRetries: int = 3
    MaxNumberOfReadRetries: int = 3
    MainFunctionPeriodMs: int = 10
    SizeStandardJobQueue: int = 16
    SizeImmediateJobQueue: int = 4

@dataclass
class NvMConfig:
    """Complete NvM Configuration"""
    general: NvMGeneralConfig
    blocks: List[NvMBlockDescriptor]

#==============================================================================
# Template for NvM_Cfg.h
#==============================================================================
NVM_CFG_H_TEMPLATE = """/**
 * @file NvM_Cfg.h
 * @brief NVRAM Manager Configuration Header (Auto-Generated)
 * @version 1.0.0
 * @date {date}
 * @author NvM Configurator Tool
 * 
 * DO NOT EDIT MANUALLY - This file is auto-generated
 */

#ifndef NVM_CFG_H
#define NVM_CFG_H

/*==================================================================================================
*                                    PRE-COMPILE CONFIGURATION
==================================================================================================*/
#define NVM_DEV_ERROR_DETECT            ({dev_error_detect})
#define NVM_VERSION_INFO_API            ({version_info_api})
#define NVM_SET_RAM_BLOCK_STATUS_API    ({set_ram_block_status_api})
#define NVM_GET_ERROR_STATUS_API        ({get_error_status_api})
#define NVM_SET_BLOCK_PROTECTION_API    ({set_block_protection_api})
#define NVM_GET_BLOCK_PROTECTION_API    ({get_block_protection_api})
#define NVM_SET_DATA_INDEX_API          ({set_data_index_api})
#define NVM_GET_DATA_INDEX_API          ({get_data_index_api})
#define NVM_CANCEL_JOB_API              ({cancel_job_api})
#define NVM_KILL_WRITE_ALL_API          ({kill_write_all_api})
#define NVM_KILL_READ_ALL_API           ({kill_read_all_api})
#define NVM_REPAIR_DAMAGED_BLOCKS_API   ({repair_damaged_blocks_api})

/*==================================================================================================
*                                    BLOCK CONFIGURATION
==================================================================================================*/
#define NVM_NUM_OF_NVRAM_BLOCKS         ({num_nvram_blocks}U)
#define NVM_NUM_OF_DATASETS             ({num_datasets}U)
#define NVM_NUM_OF_ROM_BLOCKS           ({num_rom_blocks}U)

/*==================================================================================================
*                                    BLOCK IDs
==================================================================================================*/
{block_ids}

/*==================================================================================================
*                                    BLOCK SIZES
==================================================================================================*/
{block_sizes}

/*==================================================================================================
*                                    RETRY CONFIGURATION
==================================================================================================*/
#define NVM_MAX_NUMBER_OF_WRITE_RETRIES ({max_write_retries}U)
#define NVM_MAX_NUMBER_OF_READ_RETRIES  ({max_read_retries}U)

/*==================================================================================================
*                                    CRC CONFIGURATION
==================================================================================================*/
#define NVM_CALC_RAM_BLOCK_CRC          ({calc_ram_block_crc})
#define NVM_USE_CRC_COMP_MECHANISM      ({use_crc_comp_mechanism})

/*==================================================================================================
*                                    MAIN FUNCTION PERIOD
==================================================================================================*/
#define NVM_MAIN_FUNCTION_PERIOD_MS     ({main_function_period}U)

/*==================================================================================================
*                                    QUEUE SIZES
==================================================================================================*/
#define NVM_SIZE_STANDARD_JOB_QUEUE     ({std_job_queue_size}U)
#define NVM_SIZE_IMMEDIATE_JOB_QUEUE    ({imm_job_queue_size}U)

/*==================================================================================================
*                                    MULTI BLOCK REQUESTS
==================================================================================================*/
#define NVM_MULTI_BLOCK_CALLBACK        (STD_OFF)

#endif /* NVM_CFG_H */
"""

#==============================================================================
# Template for NvM_Lcfg.c
#==============================================================================
NVM_LCFG_C_TEMPLATE = """/**
 * @file NvM_Lcfg.c
 * @brief NVRAM Manager Link-Time Configuration (Auto-Generated)
 * @version 1.0.0
 * @date {date}
 * @author NvM Configurator Tool
 * 
 * DO NOT EDIT MANUALLY - This file is auto-generated
 */

#include "NvM.h"
#include "NvM_Cfg.h"

/*==================================================================================================
*                                    ROM BLOCK DATA
==================================================================================================*/
{rom_block_data}

/*==================================================================================================
*                                    BLOCK DESCRIPTOR TABLE
==================================================================================================*/
#define NVM_START_SEC_CONFIG_DATA_UNSPECIFIED
#include "MemMap.h"

STATIC const NvM_BlockDescriptorType NvM_BlockDescriptorTable[NVM_NUM_OF_NVRAM_BLOCKS] = {{
{block_descriptors}
}};

#define NVM_STOP_SEC_CONFIG_DATA_UNSPECIFIED
#include "MemMap.h"

/*==================================================================================================
*                                    GLOBAL CONFIGURATION
==================================================================================================*/
#define NVM_START_SEC_CONFIG_DATA_UNSPECIFIED
#include "MemMap.h"

const NvM_ConfigType NvM_Config = {{
    .BlockDescriptors = NvM_BlockDescriptorTable,
    .NumBlockDescriptors = NVM_NUM_OF_NVRAM_BLOCKS,
    .NumOfNvBlocks = NVM_NUM_OF_NVRAM_BLOCKS,
    .NumOfDataSets = NVM_NUM_OF_DATASETS,
    .NumOfRomBlocks = NVM_NUM_OF_ROM_BLOCKS,
    .MaxNumberOfWriteRetries = NVM_MAX_NUMBER_OF_WRITE_RETRIES,
    .MaxNumberOfReadRetries = NVM_MAX_NUMBER_OF_READ_RETRIES,
    .DevErrorDetect = NVM_DEV_ERROR_DETECT,
    .VersionInfoApi = NVM_VERSION_INFO_API,
    .SetRamBlockStatusApi = NVM_SET_RAM_BLOCK_STATUS_API,
    .GetErrorStatusApi = NVM_GET_ERROR_STATUS_API,
    .SetBlockProtectionApi = NVM_SET_BLOCK_PROTECTION_API,
    .GetBlockProtectionApi = NVM_GET_BLOCK_PROTECTION_API,
    .SetDataIndexApi = NVM_SET_DATA_INDEX_API,
    .GetDataIndexApi = NVM_GET_DATA_INDEX_API,
    .CancelJobApi = NVM_CANCEL_JOB_API,
    .KillWriteAllApi = NVM_KILL_WRITE_ALL_API,
    .KillReadAllApi = NVM_KILL_READ_ALL_API,
    .RepairDamagedBlocksApi = NVM_REPAIR_DAMAGED_BLOCKS_API,
    .CalcRamBlockCrc = NVM_CALC_RAM_BLOCK_CRC,
    .UseCrcCompMechanism = NVM_USE_CRC_COMP_MECHANISM,
    .MainFunctionPeriod = NVM_MAIN_FUNCTION_PERIOD_MS
}};

#define NVM_STOP_SEC_CONFIG_DATA_UNSPECIFIED
#include "MemMap.h"
"""

#==============================================================================
# Generator Class
#==============================================================================
class NvMConfigGenerator:
    """Generates NvM configuration files from configuration data"""
    
    def __init__(self, config: NvMConfig):
        self.config = config
        
    def generate_cfg_h(self) -> str:
        """Generate NvM_Cfg.h content"""
        general = self.config.general
        
        # Generate block IDs
        block_ids = []
        block_sizes = []
        
        for block in self.config.blocks:
            block_ids.append(f'#define NVM_BLOCK_ID_{block.BlockName.upper()}      ((NvM_BlockIdType){block.BlockId}U)')
            block_sizes.append(f'#define NVM_BLOCK_SIZE_{block.BlockName.upper()}      ({block.NvBlockLength}U)')
        
        return NVM_CFG_H_TEMPLATE.format(
            date="2026-04-28",
            dev_error_detect=general.DevErrorDetect.value,
            version_info_api=general.VersionInfoApi.value,
            set_ram_block_status_api=general.SetRamBlockStatusApi.value,
            get_error_status_api=general.GetErrorStatusApi.value,
            set_block_protection_api=general.SetBlockProtectionApi.value,
            get_block_protection_api=general.GetBlockProtectionApi.value,
            set_data_index_api=general.SetDataIndexApi.value,
            get_data_index_api=general.GetDataIndexApi.value,
            cancel_job_api=general.CancelJobApi.value,
            kill_write_all_api=general.KillWriteAllApi.value,
            kill_read_all_api=general.KillReadAllApi.value,
            repair_damaged_blocks_api=general.RepairDamagedBlocksApi.value,
            num_nvram_blocks=len(self.config.blocks),
            num_datasets=general.SizeStandardJobQueue,
            num_rom_blocks=sum(1 for b in self.config.blocks if b.RomBlockNum > 0),
            block_ids='\n'.join(block_ids),
            block_sizes='\n'.join(block_sizes),
            max_write_retries=general.MaxNumberOfWriteRetries,
            max_read_retries=general.MaxNumberOfReadRetries,
            calc_ram_block_crc=general.CalcRamBlockCrc.value,
            use_crc_comp_mechanism=general.UseCrcCompMechanism.value,
            main_function_period=general.MainFunctionPeriodMs,
            std_job_queue_size=general.SizeStandardJobQueue,
            imm_job_queue_size=general.SizeImmediateJobQueue
        )
    
    def generate_lcfg_c(self) -> str:
        """Generate NvM_Lcfg.c content"""
        # Generate block descriptors
        descriptors = []
        rom_data = []
        
        for block in self.config.blocks:
            # ROM block data
            if block.RomBlockNum > 0:
                rom_name = f'NvM_RomBlock_{block.BlockName}'
                rom_data.append(f'static const uint8 {rom_name}[{block.NvBlockLength}] = {{0}}; /* Default values */')
            else:
                rom_name = 'NULL_PTR'
            
            # Block descriptor
            desc = f'''    {{
        .BlockId = {block.BlockId},
        .BlockBaseNumber = {block.BlockBaseNumber},
        .ManagementType = {block.ManagementType.value},
        .NumberOfNvBlocks = {block.NumberOfNvBlocks},
        .NumberOfDataSets = {block.NumberOfDataSets},
        .NvBlockLength = {block.NvBlockLength},
        .NvBlockNum = {block.NvBlockNum},
        .RomBlockNum = {block.RomBlockNum},
        .InitCallback = {block.InitCallback},
        .JobEndCallback = {block.JobEndCallback},
        .CrcType = {block.CrcType.value},
        .BlockUseCrc = {'TRUE' if block.BlockUseCrc else 'FALSE'},
        .BlockUseSetRamBlockStatus = {'TRUE' if block.BlockUseSetRamBlockStatus else 'FALSE'},
        .BlockWriteProt = {'TRUE' if block.BlockWriteProt else 'FALSE'},
        .BlockWriteOnce = {'TRUE' if block.BlockWriteOnce else 'FALSE'},
        .BlockAutoValidation = {'TRUE' if block.BlockAutoValidation else 'FALSE'},
        .BlockUseMirror = {'TRUE' if block.BlockUseMirror else 'FALSE'},
        .BlockUseCompression = {'TRUE' if block.BlockUseCompression else 'FALSE'},
        .RomBlockData = {rom_name},
        .RamBlockData = NULL_PTR,
        .MirrorBlockData = NULL_PTR
    }}'''
            descriptors.append(desc)
        
        return NVM_LCFG_C_TEMPLATE.format(
            date="2026-04-28",
            rom_block_data='\n'.join(rom_data) if rom_data else '/* No ROM blocks configured */',
            block_descriptors=','.join(descriptors)
        )
    
    def save_files(self, output_dir: Path):
        """Save generated files to output directory"""
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Save NvM_Cfg.h
        cfg_h_path = output_dir / 'NvM_Cfg.h'
        cfg_h_path.write_text(self.generate_cfg_h())
        print(f"Generated: {cfg_h_path}")
        
        # Save NvM_Lcfg.c
        lcfg_c_path = output_dir / 'NvM_Lcfg.c'
        lcfg_c_path.write_text(self.generate_lcfg_c())
        print(f"Generated: {lcfg_c_path}")

#==============================================================================
# Excel Reader
#==============================================================================
def read_excel_config(excel_path: Path) -> NvMConfig:
    """Read configuration from Excel file"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel support. Install with: pip install openpyxl")
    
    wb = openpyxl.load_workbook(excel_path)
    
    # Read general configuration
    general_sheet = wb['General']
    general = NvMGeneralConfig()
    # Parse general configuration from Excel
    
    # Read block configuration
    blocks_sheet = wb['Blocks']
    blocks = []
    
    for row in blocks_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        
        block = NvMBlockDescriptor(
            BlockId=row[0],
            BlockName=row[1],
            BlockBaseNumber=row[2] if row[2] else row[0] + 1,
            ManagementType=BlockManagementType(row[3]) if row[3] else BlockManagementType.NATIVE,
            NumberOfNvBlocks=row[4] if row[4] else 1,
            NumberOfDataSets=row[5] if row[5] else 1,
            NvBlockLength=row[6],
            NvBlockNum=row[7] if row[7] else 1,
            RomBlockNum=row[8] if row[8] else 0,
            CrcType=BlockCrcType(row[9]) if row[9] else BlockCrcType.CRC16,
            BlockUseCrc=row[10] if row[10] is not None else True,
            BlockUseSetRamBlockStatus=row[11] if row[11] is not None else True,
            BlockWriteProt=row[12] if row[12] is not None else False,
            BlockWriteOnce=row[13] if row[13] is not None else False,
            BlockAutoValidation=row[14] if row[14] is not None else False,
            BlockUseMirror=row[15] if row[15] is not None else False,
            BlockUseCompression=row[16] if row[16] is not None else False,
            Description=row[17] if len(row) > 17 else ""
        )
        blocks.append(block)
    
    return NvMConfig(general=general, blocks=blocks)

#==============================================================================
# JSON Config Reader
#==============================================================================
def read_json_config(json_path: Path) -> NvMConfig:
    """Read configuration from JSON file"""
    with open(json_path, 'r') as f:
        data = json.load(f)
    
    general = NvMGeneralConfig(**data['general'])
    blocks = [NvMBlockDescriptor(**b) for b in data['blocks']]
    
    return NvMConfig(general=general, blocks=blocks)

#==============================================================================
# Sample Configuration
#==============================================================================
def create_sample_config() -> NvMConfig:
    """Create a sample NvM configuration"""
    general = NvMGeneralConfig()
    
    blocks = [
        NvMBlockDescriptor(
            BlockId=1,
            BlockName="Config",
            BlockBaseNumber=1,
            ManagementType=BlockManagementType.NATIVE,
            NumberOfNvBlocks=1,
            NumberOfDataSets=1,
            NvBlockLength=64,
            NvBlockNum=1,
            RomBlockNum=1,
            CrcType=BlockCrcType.CRC16,
            BlockUseCrc=True,
            BlockUseSetRamBlockStatus=True,
            BlockWriteProt=False,
            BlockWriteOnce=False,
            BlockAutoValidation=True,
            BlockUseMirror=False,
            BlockUseCompression=False,
            Description="System configuration data"
        ),
        NvMBlockDescriptor(
            BlockId=2,
            BlockName="Calibration",
            BlockBaseNumber=2,
            ManagementType=BlockManagementType.DATASET,
            NumberOfNvBlocks=4,
            NumberOfDataSets=4,
            NvBlockLength=256,
            NvBlockNum=4,
            RomBlockNum=4,
            CrcType=BlockCrcType.CRC16,
            BlockUseCrc=True,
            BlockUseSetRamBlockStatus=True,
            BlockWriteProt=False,
            BlockWriteOnce=False,
            BlockAutoValidation=False,
            BlockUseMirror=True,
            BlockUseCompression=False,
            Description="Calibration parameters with dataset support"
        ),
        NvMBlockDescriptor(
            BlockId=3,
            BlockName="FaultMemory",
            BlockBaseNumber=3,
            ManagementType=BlockManagementType.REDUNDANT,
            NumberOfNvBlocks=2,
            NumberOfDataSets=1,
            NvBlockLength=512,
            NvBlockNum=2,
            RomBlockNum=0,
            CrcType=BlockCrcType.CRC32,
            BlockUseCrc=True,
            BlockUseSetRamBlockStatus=True,
            BlockWriteProt=False,
            BlockWriteOnce=False,
            BlockAutoValidation=False,
            BlockUseMirror=False,
            BlockUseCompression=False,
            Description="Fault memory with redundant storage"
        ),
        NvMBlockDescriptor(
            BlockId=4,
            BlockName="VIN",
            BlockBaseNumber=4,
            ManagementType=BlockManagementType.NATIVE,
            NumberOfNvBlocks=1,
            NumberOfDataSets=1,
            NvBlockLength=17,
            NvBlockNum=1,
            RomBlockNum=1,
            CrcType=BlockCrcType.CRC8,
            BlockUseCrc=True,
            BlockUseSetRamBlockStatus=False,
            BlockWriteProt=True,
            BlockWriteOnce=True,
            BlockAutoValidation=False,
            BlockUseMirror=False,
            BlockUseCompression=False,
            Description="Vehicle Identification Number - Write Once"
        ),
        NvMBlockDescriptor(
            BlockId=5,
            BlockName="Odometer",
            BlockBaseNumber=5,
            ManagementType=BlockManagementType.REDUNDANT,
            NumberOfNvBlocks=2,
            NumberOfDataSets=1,
            NvBlockLength=8,
            NvBlockNum=2,
            RomBlockNum=0,
            CrcType=BlockCrcType.CRC16,
            BlockUseCrc=True,
            BlockUseSetRamBlockStatus=True,
            BlockWriteProt=False,
            BlockWriteOnce=False,
            BlockAutoValidation=True,
            BlockUseMirror=True,
            BlockUseCompression=False,
            Description="Odometer reading with high reliability"
        ),
    ]
    
    return NvMConfig(general=general, blocks=blocks)

#==============================================================================
# Main Entry Point
#==============================================================================
def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='NvM Configuration Generator')
    parser.add_argument('--input', '-i', type=Path, help='Input configuration file (JSON or Excel)')
    parser.add_argument('--output', '-o', type=Path, default=Path('./output'), help='Output directory')
    parser.add_argument('--sample', '-s', action='store_true', help='Generate sample configuration')
    parser.add_argument('--format', '-f', choices=['json', 'excel'], help='Export sample as JSON or Excel')
    
    args = parser.parse_args()
    
    if args.sample:
        # Generate sample configuration
        config = create_sample_config()
        
        if args.format == 'json':
            # Export as JSON
            output_file = args.output / 'nvm_config_sample.json'
            output_file.parent.mkdir(parents=True, exist_ok=True)
            with open(output_file, 'w') as f:
                json.dump({
                    'general': asdict(config.general),
                    'blocks': [asdict(b) for b in config.blocks]
                }, f, indent=2)
            print(f"Generated sample JSON: {output_file}")
            
        elif args.format == 'excel' and OPENPYXL_AVAILABLE:
            # Export as Excel
            output_file = args.output / 'nvm_config_template.xlsx'
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            wb = openpyxl.Workbook()
            
            # General sheet
            ws_general = wb.active
            ws_general.title = 'General'
            ws_general.append(['Parameter', 'Value', 'Description'])
            general_dict = asdict(config.general)
            for key, value in general_dict.items():
                ws_general.append([key, value, ''])
            
            # Blocks sheet
            ws_blocks = wb.create_sheet('Blocks')
            ws_blocks.append([
                'BlockId', 'BlockName', 'BlockBaseNumber', 'ManagementType',
                'NumberOfNvBlocks', 'NumberOfDataSets', 'NvBlockLength',
                'NvBlockNum', 'RomBlockNum', 'CrcType', 'BlockUseCrc',
                'BlockUseSetRamBlockStatus', 'BlockWriteProt', 'BlockWriteOnce',
                'BlockAutoValidation', 'BlockUseMirror', 'BlockUseCompression',
                'Description'
            ])
            
            for block in config.blocks:
                block_dict = asdict(block)
                ws_blocks.append([
                    block_dict.get('BlockId'),
                    block_dict.get('BlockName'),
                    block_dict.get('BlockBaseNumber'),
                    block_dict.get('ManagementType'),
                    block_dict.get('NumberOfNvBlocks'),
                    block_dict.get('NumberOfDataSets'),
                    block_dict.get('NvBlockLength'),
                    block_dict.get('NvBlockNum'),
                    block_dict.get('RomBlockNum'),
                    block_dict.get('CrcType'),
                    block_dict.get('BlockUseCrc'),
                    block_dict.get('BlockUseSetRamBlockStatus'),
                    block_dict.get('BlockWriteProt'),
                    block_dict.get('BlockWriteOnce'),
                    block_dict.get('BlockAutoValidation'),
                    block_dict.get('BlockUseMirror'),
                    block_dict.get('BlockUseCompression'),
                    block_dict.get('Description')
                ])
            
            wb.save(output_file)
            print(f"Generated sample Excel: {output_file}")
        else:
            # Generate C files directly
            generator = NvMConfigGenerator(config)
            generator.save_files(args.output)
            
    elif args.input:
        # Read input and generate
        if args.input.suffix == '.json':
            config = read_json_config(args.input)
        elif args.input.suffix in ['.xlsx', '.xls']:
            config = read_excel_config(args.input)
        else:
            print(f"Unsupported file format: {args.input.suffix}")
            sys.exit(1)
        
        generator = NvMConfigGenerator(config)
        generator.save_files(args.output)
    else:
        parser.print_help()
        sys.exit(1)

if __name__ == '__main__':
    main()
